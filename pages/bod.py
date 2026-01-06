import traceback
from io import BytesIO
from datetime import date
from calendar import monthrange
import json
import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from copy import copy
from utils import json_utils
from utils import format

def viagens_expressas():
    exp = pd.read_csv(up_expressas, sep=';', encoding='Windows-1252', skiprows=2) # Pula as 2 primeiras linhas
    exp = exp.drop(columns=exp.columns[4]) # Retira a coluna sem valores

    # Define o sentido e descarta colunas antigas
    def definir_sentido(row):
        ida_local = row['Local origem'][:2]
        volta_local = row['Local destino'][:2]

        if volta_local == 'PA':
            return 1
        elif ida_local == 'PA':
            return 2
        elif ida_local == 'FA' and volta_local != 'PA':
            return 2
        elif volta_local == 'FA' and ida_local != 'PA':
            return 1
        elif ida_local == 'FL' and volta_local != 'PA':
            return 2
        elif volta_local == 'FL' and ida_local != 'PA':
            return 1
        elif ida_local == 'ES' and volta_local != 'PA':
            return 2
        elif volta_local == 'ES' and ida_local != 'PA':
            return 1
        else:
            return 1

    exp['Sentido'] = exp.apply(definir_sentido, axis=1)
    exp = (exp.groupby(['Número Linha', 'Sentido'], as_index=False)['Qt.Viagens']
        .sum()
        .sort_values(by=['Número Linha', 'Sentido'])
    )

    return exp

def dados_transnet():
    linhas = pd.read_csv(up_linhas, sep=';', encoding='Windows-1252')
    linhas = linhas.drop(linhas.columns[[2, 4, 5, 8, 9, 11, 12, 13, 15, 16, 30]], axis=1) # Retira colunas sem valores ou não usadas [EMP, R, NOME LINHA, EXTA, EXTB, TMESC, TMINT, LOTAÇÃO, KMROD, FRT]
    
    # Atribuindo o novo cabeçalho às colunas correspondentes
    linhas = linhas.rename(columns={'LINHA': 'COD', 'TS': 'SERV', 'S': 'SENT'})

    novos_nomes = ['TEU_VT', 'TEU_BIL', 'DIN', 'ESC', 'ISE', 'INT_TEU', 'INT_TAL', 'BAL', 'TEU_VT_R$', 'TEU_BIL_R$', 'DIN_R$', 'ESC_R$', 'INT_R$'] # Outras
    
    for i, nome in zip([7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19], novos_nomes):
        linhas.columns.values[i] = nome

    linhas = linhas.dropna(subset=['COD']) # Exclui as linhas NAN
    linhas = linhas.drop(index=linhas.index[-1])
    linhas['SENT'] = linhas['SENT'].replace('3', '1') # Corrige o sentido quando a linha for circular
    linhas['SERV'] = linhas['SERV'].replace('E', 'S') # Corrige o tipo de serviço de E para S no executivo
    linhas.iloc[:, 7:19] = linhas.iloc[:, 7:19].astype(str).apply(lambda col: col.str.replace('.', '', regex=False)) # Retiro o '.'
    
    # Conversões
    linhas['TMCOM'] = linhas['TMCOM'].astype(float) # float
    colunas = ['ANO','MES','SENT','TEU_VT','TEU_BIL','DIN','ESC','ISE','INT_TEU','INT_TAL','BAL'] # Converto para int
    linhas[colunas] = linhas[colunas].apply(lambda col: pd.to_numeric(col, errors='coerce')).astype(int)
    colunas = ['TEU_VT_R$','TEU_BIL_R$','DIN_R$','ESC_R$','INT_R$'] # Converto para float
    linhas[colunas] = linhas[colunas].apply(lambda col: col.str.replace('.', '', regex=False).str.replace(',', '.', regex=False)).astype(float)
    
    return linhas

def dados_PLE():
    ple = pd.read_csv(up_ple, sep=';', encoding='Windows-1252')

    # Excluir a coluna 'Unnamed: 8' se existir
    if 'Unnamed: 8' in ple.columns:
        ple = ple.drop(columns=['Unnamed: 8'])

    # Exclui colunas desnecessárias
    columns_to_drop = ['Código Operadora', 'Nome Operadora', 'Cartão', 'Estudante']
    ple = ple.drop(columns=columns_to_drop)

    # Formatar a coluna 'Valor' para numérico, substituindo vírgulas por pontos
    ple['Valor'] = ple['Valor'].str.replace(',', '.', regex=False).astype(float)

    ple = ple.rename(columns={'Data do Uso': 'Data', 'Linha': 'TEU'})
    
    # Agrupamento por código da linha e soma valor
    df_grouped = ple.groupby(['TEU'])['Valor'].agg(['sum', 'count']).reset_index()
    df_grouped.columns = ['TEU', 'Valor_sum', 'Valor_count']

    # Linhas com os códigos da bilhetagem
    df_linhas_teu = json_utils.ler_json("linhas_teu.json")

    df_grouped = pd.merge(df_grouped, df_linhas_teu, on='TEU', how='left')
    df_grouped = df_grouped[['TEU', 'MET', 'Valor_sum', 'Valor_count']]
    df_grouped = df_grouped.sort_values(by='MET').reset_index(drop=True)

    # Check for rows where 'MET' is NaN (meaning no match was found in df_json)
    missing_met = df_grouped[df_grouped['MET'].isna()]

    if not missing_met.empty:
        st.warning(f"⚠️ As seguintes linhas do arquivo do Ecitop não estão presentes no arquivo de linhas_teu.json: {missing_met}")

    return df_grouped

def matriz_bod(arq):
    # Viagens expressas
    progresso.info("Lendo viagens expressas...")
    df_exp = viagens_expressas()
    
    # Dados das linhas
    progresso.info("Lendo dados das linhas...")
    df_transnet = dados_transnet()

    # Dados PLE
    progresso.info("Lendo dados PLE's...")
    df_ple = dados_PLE()

    # Lendo a matriz BOD
    progresso.info("Lendo matriz...")
    df_matriz = pd.read_excel(arq, sheet_name='MATRIZ', decimal=',')

    # Verifica se tem alguma linha no Transnet e na matriz não
    divergencia = df_transnet[~df_transnet['COD'].isin(df_matriz['COD'])]['COD']
    if not divergencia.empty:
        st.warning(f"⚠️ As seguintes linhas do arquivo do Transnet não foram inseridas porque não foram encontradas na aba MATRIZ: {divergencia.tolist()}")

    # Verifica se tem alguma linha nos dados PLE e na matriz não
    divergencia_ple = df_ple[~df_ple['MET'].isin(df_matriz['COD'])]['COD']
    if not divergencia_ple.empty:
        st.warning(f"⚠️ As seguintes linhas dos dados PLE não foram inseridas porque não foram encontradas na aba MATRIZ: {divergencia_ple.tolist()}")

    progresso.info("Preparando matriz...")
    df_matriz["ANO"] = df_matriz["ANO"].fillna(df_transnet["ANO"].iloc[0]).astype(int) # Ano
    df_matriz["MES"] = df_matriz["MES"].fillna(df_transnet["MES"].iloc[0]).astype(int) # Mês
    
    # df_temp para merge dos dataframes
    colunas_int = df_transnet.columns[6:15] # colunas que devem virar Int64
    colunas_float = df_transnet.columns[[5, 15, 16, 17, 18, 19]] # colunas que devem virar float
    df_temp = df_transnet[['COD', 'SENT']].join(df_transnet[colunas_int.union(colunas_float)]) # Cria temp com apenas as colunas desejadas + chaves

    # Converte os tipos diretamente no df_temp
    df_temp[colunas_int] = df_temp[colunas_int].apply(lambda col: pd.to_numeric(col, errors='coerce')).astype('Int64')
    df_temp[colunas_float] = df_temp[colunas_float].apply(lambda col: pd.to_numeric(col, errors='coerce')).astype(float)
    
    # Merge
    df_matriz = pd.merge(df_matriz, df_temp, on=['COD', 'SENT'], how='left')
    df_matriz.iloc[:,34:50] = df_matriz.iloc[:,34:50].fillna(0) # Prenche com 0 as colunas novas com NaN

    # Definindo as colunas como índice e salvando a ordenação original
    ordem_colunas = df_matriz.columns.tolist()
    df_matriz.set_index(['COD', 'SENT'], inplace=True)
    df_exp.set_index(['Número Linha', 'Sentido'], inplace=True)

    # Preenchendo a coluna viagens expressas
    progresso.info("Inserindo viagens expressas...")
    df_matriz.loc[df_exp.index, 'VR_EXP'] = df_exp['Qt.Viagens']

    # ---------- Código NOVO
    # 1) Trazer a coluna C do df_ple para o df_matriz
    df_matriz = df_matriz.reset_index().merge(df_ple[["MET", "Valor_count", "Valor_sum"]], on="COD", how="left")

    # 2) Calcular o total de D por A
    df_matriz["peso_total"] = df_matriz.groupby("COD")["PASS_COM"].transform("sum")

    # 3) Proporção de cada linha
    df_matriz["proporcao"] = df_matriz["PASS_COM"] / df_matriz["peso_total"]

    # 4) Preencher a coluna E existente com o rateio
    df_matriz["PASS_LIVRE"] = df_matriz["Valor_count"] * df_matriz["proporcao"]
    df_matriz["REC_TAR_LIVRE"] = df_matriz["Valor_sum"] * df_matriz["proporcao"]

    # 5) Restaurar o índice original (A, B)
    df_matriz = df_matriz.set_index(["COD", "SENT"])

    # 6) Remover colunas auxiliares
    df_matriz = df_matriz.drop(columns=["peso_total", "proporcao"])


    # Preenchendo colunas
    progresso.info("Inserindo informações...")
    df_matriz.loc[df_matriz['TAR_MAX_COM'] == 0, 'TAR_MAX_COM'] = df_matriz['TMCOM'] # Tárifa máxima comum
    df_matriz["TAR_MAX_COM"] = df_matriz["TAR_MAX_COM"].fillna(0)
    df_matriz["TAR_MAX_ESC"] = (df_matriz["TAR_MAX_COM"] * 0.9).where(df_matriz["SERV"] != "S") # Tárifa máxima escolar
    df_matriz["TAR_MAX_ESC"] = df_matriz["TAR_MAX_ESC"].fillna(df_matriz["TAR_MAX_COM"])
    df_matriz["EXTP_SIMP"] = df_matriz["KM"] * df_matriz["VGR"] # Extensão simples
    df_matriz["EXTP_EXP"] = df_matriz["KM_EXP"] * df_matriz["VR_EXP"] # Extensão expressa
    df_matriz["VR_SIMP"] = df_matriz["VGR"] # Viagens normais
    df_matriz["PASS_COM"] = df_matriz['TEU_VT'] + df_matriz['TEU_BIL'] + df_matriz['DIN'] # Passageiro comum
    df_matriz["PASS_ESC"] = df_matriz['ESC'] # Passageiro escolar
    #df_matriz["PASS_LIVRE"] = 1 # Passageiro PLE (passe livre estudantil)
    df_matriz["PASS_ISE"] = (df_matriz['ISE'] + (df_matriz['TX_ISE'] * (df_matriz['TEU_VT'] + df_matriz['TEU_BIL'] + df_matriz['DIN'] + df_matriz['ESC'] + df_matriz['INT_TEU'] + df_matriz['INT_TAL'] + df_matriz['BAL']))).round() # Passageiro isento
    df_matriz["PASS_INT_ROD"] = df_matriz['INT_TEU'] + df_matriz['INT_TAL'] # Passageiro int rodov
    df_matriz["REC_TAR_COM"] = df_matriz['TEU_VT_R$'] + df_matriz['TEU_BIL_R$'] + df_matriz['DIN_R$'] + df_matriz['INT_R$'] # Receita tarifa comum
    df_matriz["REC_TAR_ESC"] = df_matriz['ESC_R$'] # Receita tarifa escolar
    #df_matriz["REC_TAR_LIVRE"] = df_matriz['ESC_R$'] # Receita tarifa PLE (passe livre estudantil)
    
    # Colunas/variáveis temporárias para cálculos das próximas colunas
    progresso.info("Calculando colunas...")
    df_matriz['KM_LINHA'] = df_matriz['EXTP_SIMP'] + df_matriz['EXTP_EXP']
    comb_real = df_matriz['KM_LINHA'].sum() # Soma o KM_LINHA para obter o combustível real
    comb_desloc = km - comb_real # Combustível deslocamento, KM informado na página - combustível real
    df_matriz['KM_PROP'] = df_matriz['KM_LINHA'].apply(lambda x: 0 if x == 0 else x / comb_real) # KM proporcional, se KM_LINHA não for 0 então KM_LINHA / combustível real

    # Colunas normais que necessitam dos cálculos acima
    df_matriz['EXTP_DESL'] = round(df_matriz['KM_PROP'] * comb_desloc, 2)
    df_matriz['VR_DESL'] = round(df_matriz['EXTP_DESL'] / df_matriz['KM_OSC'])

    df_matriz.reset_index(inplace=True) # Reseto os índices
    df_matriz = df_matriz[ordem_colunas] # Reorganizo as colunas

    df_exp = None
    df_transnet = None
    df_temp = None

    return df_matriz

# Configuração da página
st.set_page_config(layout="wide")

# Lê arquivo de configuração
config = json_utils.ler_json("config.json")

st.info(fr"Último período gerado: {config['bod']['periodo']}", icon="ℹ️")

# Colunas
col1, col2, col3, col4 = st.columns([2, 2, 2, 1], vertical_alignment='top')

with col1:
    # Upload do arquivo de viagens expressas
    st.subheader("Viagens expressas")
    up_expressas = st.file_uploader("Selecione um arquivo .CSV", type='csv', key=1)
     
with col2:
    # Upload do arquivo dos dados das linhas
    st.subheader("Dados das linhas")
    up_linhas = st.file_uploader("Selecione um arquivo .CSV", type='csv', key=2)

with col3:
    # Upload do arquivo PLE
    st.subheader("PLE")
    up_ple = st.file_uploader("Selecione um arquivo .CSV", type='csv', key=3)    

with col4:
    # KM mensal
    st.subheader("KM Mensal")
    km = st.number_input("KM", value=0)


botao = st.sidebar.button("Iniciar", type="primary")

st.divider()
    
if botao:
    try:
        # Verificações e leitura dos arquivos
        if up_expressas is None:
            st.warning("Arquivo de viagens expressas não foi selecionado!", icon=":material/error_outline:")
            st.stop()

        if up_linhas is None:
            st.warning("Arquivo de dados das linhas não foi selecionado!", icon=":material/error_outline:")        
            st.stop()

        if up_ple is None:
            st.warning("Arquivo PLE não foi selecionado!", icon=":material/error_outline:")
            st.stop()

        if (km is None) or (km <= 0):
            st.warning("KM mensal não está de acordo!", icon=":material/error_outline:")
            st.stop()

        # Processando
        progresso = st.empty()
        
        df_bod = matriz_bod(config['bod']['modelo_bod'])
        mes, ano = df_bod.iloc[1][['MES', 'ANO']] # Pego mês e ano para os nomes dos arquivos

        # ✳️ Preencher planilha BOD Metroplan ✳️
        progresso.info("Preenchendo planilha Metroplan...")

        wb_met = load_workbook(config['bod']['modelo_metroplan'])
        wb_met['Identificação da Empresa'].cell(row=11, column=7, value=mes)
        wb_met['Identificação da Empresa'].cell(row=11, column=8, value=ano)
        ws = wb_met['BOD']
        ws.protection.sheet = False
        ws.protection.disable()

        num_linhas_df = len(df_bod)

        # Escrever os dados
        for i, row in enumerate(df_bod.itertuples(index=False), start=2):
            for j, valor in enumerate(row[:29], start=1): # Até a coluna Frota
                ws.cell(row=i, column=j, value=valor)

        # Definir formatação pois não pegou a formatação da planilha
        for i in range(2, num_linhas_df + 2):
            for col in range(9, 11):  
                ws.cell(row=i, column=col).number_format = '0.000'
            for col in range(11, 14):  
                ws.cell(row=i, column=col).number_format = '0.00'
            for col in range(14, 26):  
                ws.cell(row=i, column=col).number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)' # Contábil sem sinal e casas decimais
            for col in range(26, 29):  
                ws.cell(row=i, column=col).number_format = '0.00'

        progresso.info("Salvando BOD Metroplan...")

        # Salva em memória 
        buffer_met = BytesIO() 
        wb_met.save(buffer_met) 
        buffer_met.seek(0)
        
        ws = None
        wb_met = None

        # ✳️ Preencher planilha BOD ERG ✳️
        wb_erg = load_workbook(config['modelo_bod'])

        # 1️⃣ [BOD]
        progresso.info("Preenchendo planilha BOD [BOD]...")
        ws_bod = wb_erg['BOD']
        linha_modelo = 3 # Linha com a formatação de referência
        num_linhas_df = len(df_bod)

        # Copiar formatação
        for i in range(num_linhas_df):
            for col in range(1, 30):  
                celula_origem = ws_bod.cell(row=linha_modelo, column=col)
                celula_destino = ws_bod.cell(row=linha_modelo + i, column=col)
                celula_destino._style = copy(celula_origem._style)

        # Escrever os dados
        for i, row in enumerate(df_bod.itertuples(index=False), start=linha_modelo):
            for j, valor in enumerate(row[:29], start=1): # Até a coluna Frota
                ws_bod.cell(row=i, column=j, value=valor)

        #Salvo os dados no arquivo de config para o PDO
        progresso.info("Atualizando arquivo de config com os dados do BOD...")

        config['bod']['periodo'] = fr"{mes:02}/{ano}"
        config['bod']['bod_km_linhas_1'] = df_bod.loc[(df_bod['SENT'] == 1) & (df_bod['COD'] != 'M105'), 'EXTP_SIMP'].sum()
        config['bod']['bod_km_linhas_2'] = df_bod.loc[(df_bod['SENT'] == 2) & (df_bod['COD'] != 'M105'), 'EXTP_SIMP'].sum()
        config['bod']['bod_km_tm5_1'] = df_bod.loc[(df_bod['SENT'] == 1) & (df_bod['COD'] == 'M105'), 'EXTP_SIMP'].sum()
        config['bod']['bod_km_tm5_2'] = df_bod.loc[(df_bod['SENT'] == 2) & (df_bod['COD'] == 'M105'), 'EXTP_SIMP'].sum()
        config['bod']['bod_isentos_linhas_1'] = df_bod.loc[(df_bod['SENT'] == 1) & (df_bod['COD'] != 'M105'), 'PASS_ISE'].sum()
        config['bod']['bod_isentos_linhas_2'] = df_bod.loc[(df_bod['SENT'] == 2) & (df_bod['COD'] != 'M105'), 'PASS_ISE'].sum()
        config['bod']['bod_isentos_tm5_1'] = df_bod.loc[(df_bod['SENT'] == 1) & (df_bod['COD'] == 'M105'), 'PASS_ISE'].sum()
        config['bod']['bod_isentos_tm5_2'] = df_bod.loc[(df_bod['SENT'] == 2) & (df_bod['COD'] == 'M105'), 'PASS_ISE'].sum()

        with open('config.json', 'w', encoding='utf-8') as arq_conf:
            json.dump(config, arq_conf, indent=2, ensure_ascii=False)

        # 2️⃣ [SINTETICO]
        progresso.info("Preenchendo planilha BOD [SINTETICO]...")
        # Montagem do dataframe. Aproveito e incluo as colunas para serem usadas na aba ATM
        # Mapeamento das colunas e renomeando para os novos nomes desejados
        colunas_renomeadas = {
            'VR_SIMP': 'VU',
            'VR_EXP': 'VE',
            'EXTP_SIMP': 'RU',
            'EXTP_EXP': 'RE',
            'BAL': 'B',
            'INT_TEU': 'INT_TEU',
            'INT_TAL': 'INT_TAL',
            'ISE': 'I',
            'TEU_VT': 'TEU_VT',
            'TEU_BIL': 'TEU_BIL',
            'ESC': 'PE',
            'DIN': 'PL',
            'DIN_R$': 'R',
            'PASS_ISE': 'ISENTOS', # Colunas para a aba ATM
            'REC_TAR_COM': 'REC_TAR_COM',
            'REC_TAR_ESC': 'REC_TAR_ESC'
        }

        # Agrupamento com renomeação já no resultado
        df_agrupado = df_bod.groupby('COD').agg({col: 'sum' for col in colunas_renomeadas}).rename(columns=colunas_renomeadas).reset_index()

        # Cálculos derivados
        df_agrupado['INT'] = df_agrupado['INT_TEU'] + df_agrupado['INT_TAL']
        df_agrupado['VT'] = df_agrupado['TEU_VT'] + df_agrupado['TEU_BIL']
        df_agrupado['PP'] = df_agrupado['VT'] + df_agrupado['PE'] + df_agrupado['PL']
        df_agrupado['PT'] = df_agrupado['I'] + df_agrupado['PP']
        df_agrupado['PB'] = df_agrupado['B'] + df_agrupado['INT'] + df_agrupado['PT']
        df_agrupado['VL'] = df_agrupado['VU'] + df_agrupado['VE']
        df_agrupado['RL'] = df_agrupado['RU'] + df_agrupado['RE']
        df_agrupado['TOTAL'] = df_agrupado['VT'] + df_agrupado['PL'] + df_agrupado['PE'] + df_agrupado['ISENTOS'] # Colunas para a aba ATM
        df_agrupado['RECEITA'] = df_agrupado['REC_TAR_COM'] + df_agrupado['REC_TAR_ESC']

        # Condicionais
        df_agrupado['VEVL'] = np.where(df_agrupado['VE'] == 0, 0, df_agrupado['VE'] / df_agrupado['VL'])
        df_agrupado['VTPT'] = np.where(df_agrupado['PT'] == 0, 0, df_agrupado['VT'] / df_agrupado['PT'])
        df_agrupado['PTVL'] = np.where(df_agrupado['VL'] == 0, 0, df_agrupado['PT'] / df_agrupado['VL'])
        df_agrupado['IPK'] = np.where(df_agrupado['RL'] == 0, 0, df_agrupado['PT'] / df_agrupado['RL'])

        df_fixos = df_bod[['COD', 'NOME', 'COD_VT', 'TAR_MAX_COM']].drop_duplicates(subset='COD')
        df_sintetico = pd.merge(df_fixos, df_agrupado, on='COD', how='left')
        colunas_ordenadas = [
            'COD', 'NOME', 'COD_VT', 'VU', 'VE', 'VL', 'PB', 'B', 'INT', 'PT', 'I', 'PP', 'VT', 'PE', 'PL', 
            'R', 'RU', 'RE', 'RL', 'VEVL', 'VTPT', 'PTVL', 'IPK', 'TAR_MAX_COM', 'ISENTOS', 'TOTAL', 'RECEITA'
        ]

        df_sintetico = df_sintetico[colunas_ordenadas]

        ws_sin = wb_erg['SINTETICO']

        dados_cod = df_sintetico.set_index(['COD']).to_dict('index')  # Cria dicionário com os dados do df_final, cada COD mapeia um dict com os valores
        cods_restantes = set(dados_cod.keys()) # Lista de CODs ainda não usados
        linha = 2  # começa da linha 2

        while True:
            celula_cod = ws_sin.cell(row=linha, column=1).value

            # Verifica fim da planilha ou célula vazia
            if celula_cod is None:
                linha += 1
                continue

            # Se encontrar o texto 'TOTAL', interrompe
            if str(celula_cod).strip().upper() == 'TOTAL':
                break

            # Tenta encontrar o COD no df_sin
            if celula_cod in dados_cod:
                valores = list(dados_cod[celula_cod].values())
                for i in range(0, 22): # Vai até IPK
                    ws_sin.cell(row=linha, column=i + 2, value=valores[i])

                cods_restantes.discard(celula_cod)  # Remove o cod já inserido

            linha += 1

        # Após sair do loop, verifica se há dados não inseridos
        if cods_restantes:
            st.warning(f"⚠️ As seguintes linhas não foram inseridas porque não foram encontradas na planilha modelo [SINTETICO]: {list(cods_restantes)}")

        # 3️⃣ [ATM]
        ws_atm = wb_erg['ATM']

        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, monthrange(ano, mes)[1])
        ws_atm.cell(row=4, column=7, value=f'{primeiro_dia.strftime("%d/%m/%Y")} a {ultimo_dia.strftime("%d/%m/%Y")}') # Período
        ws_atm.cell(row=5, column=7, value=date.today().strftime("%d/%m/%Y")) # Data

        # Excluo as colunas não usadas para facilitar a iteração e ordeno
        df_sintetico = df_sintetico.drop(columns=['NOME', 'COD_VT', 'VU', 'VE', 'VL', 'PB', 'B', 'INT', 'PT', 'I', 'PP', 'R', 'RU', 'RE', 'RL', 'VEVL', 'VTPT', 'PTVL', 'IPK'])
        colunas_ordenadas = ['COD', 'TAR_MAX_COM', 'VT', 'PL', 'PE', 'ISENTOS', 'TOTAL', 'RECEITA']

        df_sintetico = df_sintetico[colunas_ordenadas]

        dados_cod = df_sintetico.set_index(['COD']).to_dict('index')
        cods_restantes = set(dados_cod.keys()) # monto novamente a lista de CODs ainda não usados

        linha = 9 

        while True:
            celula_cod = ws_atm.cell(row=linha, column=1).value.split('-')[0].strip()

            # Se encontrar o texto 'TOTAL', soma e interrompe
            if str(celula_cod).strip().upper() == 'TOTAL':
                df_soma = df_sintetico[['VT', 'PL', 'PE', 'ISENTOS', 'TOTAL', 'RECEITA']].sum()
                for indice, valor in enumerate(df_soma, start=3):
                    ws_atm.cell(row=linha, column=indice, value=valor)

                break
            
            # Tenta encontrar o COD no df_sin
            if celula_cod in dados_cod:
                for idx_coluna, (coluna, valor) in enumerate(dados_cod[celula_cod].items(), start=0):
                    ws_atm.cell(row=linha, column=idx_coluna + 2, value=valor)
                    if (idx_coluna > 0): ws_atm.cell(row=linha + 1, column=idx_coluna + 2, value=valor)

                cods_restantes.discard(celula_cod)  # Remove o cod já inserido

            linha += 2

        # Após sair do loop, verifica se há dados não inseridos
        if cods_restantes:
            st.warning(f"⚠️ As seguintes linhas não foram inseridas porque não foram encontradas na planilha modelo [ATM]: {list(cods_restantes)}")

        # Salvar
        progresso.info("Salvando BOD ERG...")
        wb_erg.remove(wb_erg["MATRIZ"])
        wb_erg.active = wb_erg.sheetnames.index('BOD')
        for aba in wb_erg.worksheets:
            aba.sheet_view.tabSelected = (aba.title == 'BOD')
        
        # Salva em memória 
        buffer_erg = BytesIO() 
        wb_erg.save(buffer_erg)
        buffer_erg.seek(0)

        # ✳️ Downloads ✳️
        st.sidebar.download_button( 
            label="Baixar BOD Metroplan", 
            data=buffer_met, 
            file_name= fr"90348517000169-BOD-TMA-{ano}{mes:02}-{ano}{mes:02}.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
            )

        st.sidebar.download_button( 
            label="Baixar BOD ERG", 
            data=buffer_erg, 
            file_name= fr"BOD {mes:02}.{ano}.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
            )


        # ✳️ Comparativos de valores ✳️
        df_soma['VT'] = df_soma['VT'] + df_soma['PL']
        df_soma = df_soma.drop(['PL', 'TOTAL'])
        df_bod_total = df_bod[['PASS_COM', 'PASS_ESC', 'PASS_ISE']].sum()
        df_bod_total['RECEITA'] = df_bod['REC_TAR_COM'].sum() + df_bod['REC_TAR_ESC'].sum()
        df_bod_total = df_bod_total.rename(index={'PASS_COM':'VT', 'PASS_ESC':'PE', 'PASS_ISE':'ISENTOS'})

        df_soma = df_soma.astype('object')
        df_bod_total = df_bod_total.astype('object')
        
        # Monta tabela comparativa
        tabela = pd.DataFrame({
             'Total BOD': pd.Series([format.formatar_valor(df_bod_total[i], moeda=(i == 'RECEITA')) for i in df_bod_total.index], index=df_bod_total.index),
             'Total ATM': pd.Series([format.formatar_valor(df_soma[i], moeda=(i == 'RECEITA')) for i in df_soma.index], index=df_soma.index)
        })

        # Adiciona a coluna de verificação
        tabela['-'] = [
            "✅" if round(df_bod_total[col], 2) == round(df_soma[col], 2) else "❌"
                for col in df_bod_total.index
        ]

        comp1, comp2 = st.columns(2)
        with comp2:
            # Estilo para alinhar à direita
            estilo = tabela.style.set_properties(**{'text-align': 'right'}).set_table_styles([
                {'selector': 'th', 'props': [('text-align', 'right')]}
            ])

            # Exibir como HTML
            st.markdown(estilo.to_html(), unsafe_allow_html=True)

        progresso.success("Processo concluído!")

    except Exception as e:    
        st.error(f"🐞 Erro: {traceback.format_exc()}")

    finally:
        wb_met = None
        wb_erg = None
        ws = None
        ws_bod = None
        ws_sin = None
        ws_atm = None
        df_bod = None
        df_agrupado = None
        df_filtrado = None
        df_fixos = None
        df_sintetico = None
        df_soma = None
        df_bod_total = None
        tabela = None
        up_expressas = None
        up_linhas = None
        progress_text = None    



