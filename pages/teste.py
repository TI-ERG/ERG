import streamlit as st
from openpyxl import load_workbook
from copy import copy
from io import BytesIO
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

st.title("Copiar Range A5:AT6 e Colar em Outro Local")

def copiar_range(ws, min_row, max_row, min_col, max_col):
    dados = []
    merges = []

    # Copiar valores e estilos
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:

            # Ignorar células mescladas falsas
            if isinstance(cell, MergedCell):
                continue

            dados.append({
                "row_offset": cell.row - min_row,
                "col_offset": cell.col_idx - min_col,
                "value": cell.value,
                "font": copy(cell.font),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
                "number_format": copy(cell.number_format),
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment)
            })

    # Copiar mesclagens internas ao bloco
    for merged in ws.merged_cells.ranges:
        if (merged.min_row >= min_row and merged.max_row <= max_row and
            merged.min_col >= min_col and merged.max_col <= max_col):

            merges.append({
                "min_row": merged.min_row - min_row,
                "max_row": merged.max_row - min_row,
                "min_col": merged.min_col - min_col,
                "max_col": merged.max_col - min_col
            })

    return {
        "dados": dados,
        "merges": merges,
        "altura": max_row - min_row + 1,
        "largura": max_col - min_col + 1
    }


def colar_range(ws, pacote, destino_row, destino_col):

    # 1. Colar valores e estilos
    for item in pacote["dados"]:
        new_row = destino_row + item["row_offset"]
        new_col = destino_col + item["col_offset"]

        cell = ws.cell(row=new_row, column=new_col)
        cell.value = item["value"]
        cell.font = item["font"]
        cell.border = item["border"]
        cell.fill = item["fill"]
        cell.number_format = item["number_format"]
        cell.protection = item["protection"]
        cell.alignment = item["alignment"]

    # 2. Colar mesclagens
    for m in pacote["merges"]:
        min_row = destino_row + m["min_row"]
        max_row = destino_row + m["max_row"]
        min_col = destino_col + m["min_col"]
        max_col = destino_col + m["max_col"]

        range_str = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

        ws.merge_cells(range_str)


uploaded_file = st.file_uploader("Selecione um arquivo .xlsx", type=["xlsx"])

if uploaded_file:
    destino_row = st.number_input("Linha de destino", min_value=1, value=20)
    destino_col = st.number_input("Coluna de destino (1 = A)", min_value=1, value=1)

    if st.button("Copiar Range A5:AT6 para o destino"):
        wb = load_workbook(uploaded_file)
        ws = wb.active

        pacote = copiar_range(ws, 16, 24, 1, 9)  # A5:AT6
        colar_range(ws, pacote, destino_row, destino_col)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Bloco copiado com sucesso!")

        st.download_button(
            label="Baixar arquivo modificado",
            data=output,
            file_name="planilha_modificada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )