import traceback
from io import BytesIO
import streamlit as st
import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell
from utils import json_utils
from utils import files_utils
from utils import date_utils

MAPA_MODAL = {"IO": "Alimentador", "CM": "Comum", "SD": "Semi-Direto"}

# region FUNÇÕES
def copiar_range(ws, min_row, max_row, min_col, max_col, values=True):
    dados = []
    merges = []

    # Copiar valores e estilos
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:

            # Ignorar células mescladas falsas
            if isinstance(cell, MergedCell):
                continue

            dados.append({
                "row_offset": cell.row - min_row,
                "col_offset": cell.col_idx - min_col,
                "value": cell.value if values else None,
                "font": copy(cell.font),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
                "number_format": copy(cell.number_format),
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment)
            })

    # Copiar mesclagens internas ao bloco
    for merged in ws.merged_cells.ranges:
        if (merged.min_row >= min_row and merged.max_row <= max_row and
            merged.min_col >= min_col and merged.max_col <= max_col):

            merges.append({
                "min_row": merged.min_row - min_row,
                "max_row": merged.max_row - min_row,
                "min_col": merged.min_col - min_col,
                "max_col": merged.max_col - min_col
            })

    return {
        "dados": dados,
        "merges": merges,
        "altura": max_row - min_row + 1,
        "largura": max_col - min_col + 1
    }

def colar_range(ws, pacote, destino_row, destino_col, values=True):
    altura = pacote["altura"]
    largura = pacote["largura"]

    # 1) Desmesclar qualquer merge que toque a área de destino
    min_row_dest = destino_row
    max_row_dest = destino_row + altura - 1
    min_col_dest = destino_col
    max_col_dest = destino_col + largura - 1

    for merge in list(ws.merged_cells.ranges):
        if (
            merge.min_row <= max_row_dest and
            merge.max_row >= min_row_dest and
            merge.min_col <= max_col_dest and
            merge.max_col >= min_col_dest
        ):
            ws.unmerge_cells(str(merge))

    # 2) Colar valores e estilos
    for item in pacote["dados"]:
        new_row = destino_row + item["row_offset"]
        new_col = destino_col + item["col_offset"]

        cell = ws.cell(row=new_row, column=new_col)

        # Se por algum motivo ainda cair em MergedCell, pula
        if isinstance(cell, MergedCell):
            continue

        if values: cell.value = item["value"]
        cell.font = item["font"]
        cell.border = item["border"]
        cell.fill = item["fill"]
        cell.number_format = item["number_format"]
        cell.protection = item["protection"]
        cell.alignment = item["alignment"]

    # 3) Recriar mesclagens relativas ao pacote
    for m in pacote["merges"]:
        min_row = destino_row + m["min_row"]
        max_row = destino_row + m["max_row"]
        min_col = destino_col + m["min_col"]
        max_col = destino_col + m["max_col"]

        range_str = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

        ws.merge_cells(range_str)

def preencher_totalizador(nova_aba, linha_inicio, linha_fim, linha_totalizador, colunas_dias):
    def to_num(v):
        if v is None or v == "":
            return 0
        if isinstance(v, (int, float)):
            return v
        # tenta converter string numérica
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return 0

    for dia_semana, col_base in colunas_dias.items():

        # Horários oficiais (contagem)
        total_horarios = sum(
            1 for r in range(linha_inicio, linha_fim + 1)
            if nova_aba.cell(row=r, column=col_base).value not in (None, "")
        )

        # Demanda oficial
        total_demanda = sum(
            to_num(nova_aba.cell(row=r, column=col_base + 1).value)
            for r in range(linha_inicio, linha_fim + 1)
        )

        # Lugares oficiais
        total_lugares = sum(
            to_num(nova_aba.cell(row=r, column=col_base + 2).value)
            for r in range(linha_inicio, linha_fim + 1)
        )

        # Horários eventuais (contagem)
        total_horarios_ev = sum(
            1 for r in range(linha_inicio, linha_fim + 1)
            if nova_aba.cell(row=r, column=col_base + 3).value not in (None, "")
        )

        # Demanda eventual
        total_demanda_ev = sum(
            to_num(nova_aba.cell(row=r, column=col_base + 4).value)
            for r in range(linha_inicio, linha_fim + 1)
        )

        # Lugares eventuais
        total_lugares_ev = sum(
            to_num(nova_aba.cell(row=r, column=col_base + 5).value)
            for r in range(linha_inicio, linha_fim + 1)
        )

        nova_aba.cell(row=linha_totalizador, column=col_base).value = total_horarios
        nova_aba.cell(row=linha_totalizador, column=col_base + 1).value = total_demanda
        nova_aba.cell(row=linha_totalizador, column=col_base + 2).value = total_lugares
        nova_aba.cell(row=linha_totalizador, column=col_base + 3).value = total_horarios_ev
        nova_aba.cell(row=linha_totalizador, column=col_base + 4).value = total_demanda_ev
        nova_aba.cell(row=linha_totalizador, column=col_base + 5).value = total_lugares_ev

def preencher_conferencia(wb, df, tm5=False):
    ws = wb["Conferência"]
    linha_excel = 4
    df_filtrado = df[df["Codigo"] == "M105"] if tm5 else df[~df["Codigo"].isin(["M105"])]

    for row in df_filtrado.itertuples(index=False): 
        ws[f"A{linha_excel}"] = row.Codigo 
        ws[f"B{linha_excel}"] = row.ERG_U1
        ws[f"C{linha_excel}"] = row.MET_U1
        ws[f"E{linha_excel}"] = row.EXT_U1
        ws[f"F{linha_excel}"] = -row.FURO_U1
        ws[f"H{linha_excel}"] = row.ERG_U2
        ws[f"I{linha_excel}"] = row.MET_U2
        ws[f"K{linha_excel}"] = row.EXT_U2
        ws[f"L{linha_excel}"] = -row.FURO_U2
        ws[f"N{linha_excel}"] = row.ERG_S1
        ws[f"O{linha_excel}"] = row.MET_S1
        ws[f"Q{linha_excel}"] = row.EXT_S1
        ws[f"R{linha_excel}"] = -row.FURO_S1
        ws[f"T{linha_excel}"] = row.ERG_S2
        ws[f"U{linha_excel}"] = row.MET_S2
        ws[f"W{linha_excel}"] = row.EXT_S2
        ws[f"X{linha_excel}"] = -row.FURO_S2
        ws[f"Z{linha_excel}"] = row.ERG_D1
        ws[f"AA{linha_excel}"] = row.MET_D1
        ws[f"AC{linha_excel}"] = row.EXT_D1
        ws[f"AD{linha_excel}"] = -row.FURO_D1
        ws[f"AF{linha_excel}"] = row.ERG_D2
        ws[f"AG{linha_excel}"] = row.MET_D2
        ws[f"AI{linha_excel}"] = row.EXT_D2
        ws[f"AJ{linha_excel}"] = -row.FURO_D2
        linha_excel += 1
    
    return wb

def preencher_totais(wb, tm5=False):
    ws = wb["Total Geral"]
    ws["D5"] = config["bod"]["bod_km_tm5_1"] if tm5 else config["bod"]["bod_km_linhas_1"]
    ws["D6"] = config["bod"]["bod_km_tm5_2"] if tm5 else config["bod"]["bod_km_linhas_2"]
    ws["D7"] = (ws["D5"].value or 0) + (ws["D6"].value or 0)

    return wb

def criar_abas_com_dias(wb, tm5=False):
    if "Modelo" not in wb.sheetnames:
        placeholder.warning("⚠️ A aba Modelo não existe na planilha.")
        st.stop()

    df_det3 = df_det2.copy()
    df_det3["Dia"] = pd.to_datetime(df_det3["Dia"], dayfirst=True, errors="coerce")
    df_filtrado = df_det3[df_det3["Codigo"] == "M105"] if tm5 else df_det3[~df_det3["Codigo"].isin(["M105"])]
    df_filtrado = df_filtrado.reset_index(drop=True)
    data_ref = df_filtrado.loc[0, "Dia"]

    aba_modelo = wb["Modelo"]
    aba_modelo["A2"] = "Nome da Empresa: Expresso Rio Guaíba"
    aba_modelo["D2"] = "Código da Empresa: GU99"
    aba_modelo["G2"] = f"Mês de referência: {data_ref.month_name(locale='pt_BR')}/{data_ref.year}"

    total_semanas = date_utils.semanas_no_mes(data_ref)

    posicoes_dias = {
        0: "E5",   # segunda
        1: "K5",   # terça
        2: "Q5",   # quarta
        3: "W5",   # quinta
        4: "AC5",  # sexta
        5: "AI5",  # sábado
        6: "AO5",  # domingo
    }

    colunas_dias = {
        d: column_index_from_string(cel.split("5")[0])
        for d, cel in posicoes_dias.items()
    }

    dias_mes = date_utils.dias_do_mes(data_ref) # Lista de datas do mês

    feriados_dict = {
        row["data"].strftime("%d/%m/%Y"): row["escala"]
        for _, row in df_feriado_editado.iterrows()
    }

    fill_feriado = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo

    for semana_num in range(1, total_semanas + 1):
        nome_aba = date_utils.semana_extenso_numero(semana_num)

        if nome_aba in wb.sheetnames:
            del wb[nome_aba]

        nova_aba = wb.copy_worksheet(aba_modelo)
        nova_aba.title = nome_aba

        # Lista de datas da semana
        dias_semana = dias_mes[dias_mes["Dia"].apply(date_utils.semana_do_mes) == semana_num]

        for _, row in dias_semana.iterrows():
            dia = row["Dia"]
            dia_semana = dia.weekday()
            cel = posicoes_dias[dia_semana]

            texto = f"{date_utils.dia_da_semana(dia)} Dia: {dia.day}"

            dia_key = dia.strftime("%d/%m/%Y")
            if dia_key in feriados_dict:
                texto += f" (Escala de {feriados_dict[dia_key]})"
                nova_aba[cel].fill = fill_feriado

            nova_aba[cel] = texto

        # Insere os dados por semana
        inserir_dados_por_semana(nova_aba, semana_num, df_filtrado, colunas_dias, MAPA_MODAL)
    
    del wb["Modelo"]
    return wb

def inserir_dados_por_semana(nova_aba, semana_num, df_det2, colunas_dias, MAPA_MODAL):
    # === 1) Copiar ranges fixos ===
    cabecalho = copiar_range(nova_aba, 5, 6, 1, 46, True)        # A5:AT6
    linha_dados = copiar_range(nova_aba, 7, 7, 1, 46, False)     # A7:AT7
    totalizador_tpl = copiar_range(nova_aba, 8, 8, 1, 46, False) # A8:AT8
    legenda = copiar_range(nova_aba, 10, 18, 1, 9, True)         # A10:I18

    # === 2) Limpar área de dados (linhas 8 até 19) ===
    for merge in list(nova_aba.merged_cells.ranges): # Remover mesclagens que tocam as linhas 8 a 19
        if merge.min_row <= 19 and merge.max_row >= 8:
            nova_aba.unmerge_cells(str(merge))

    nova_aba.delete_rows(8, 12)  # remove linhas 8 a 19 (12 linhas)
    # Ordena por Modal
    ordem_modal = list(MAPA_MODAL.keys())  # ["IO", "CM", "SD"]
    df_det2["Modal"] = pd.Categorical(df_det2["Modal"], categories=ordem_modal, ordered=True)

    # === 3) Criar df_semana ===
    df_semana = df_det2[
        (df_det2["Dia"].apply(date_utils.semana_do_mes) == semana_num) &
        (df_det2["Observacao"] != "Furo de Viagem")
    ].copy()

    df_semana["Dia"] = pd.to_datetime(df_semana["Dia"], dayfirst=True, errors="coerce")

    df_sent1 = df_semana[df_semana["Sent"] == 1].sort_values(["Modal", "Codigo", "Dia", "THor"])
    df_sent2 = df_semana[df_semana["Sent"] == 2].sort_values(["Modal", "Codigo", "Dia", "THor"])
    
    # === Barra de progresso ===
    total_blocos = df_sent1["Codigo"].nunique() + df_sent2["Codigo"].nunique()
    contador = 0

    # === 4) BLOCO SENT 1 ===
    linha_global_1 = 7 # primeira linha de dados de Sent 1

    for (codigo, sent), df_bloco in df_sent1.groupby(["Codigo", "Sent"], sort=False):
        df_bloco = df_bloco.copy()
        df_bloco["ordem"] = df_bloco.groupby("Dia").cumcount()
        max_ordem = int(df_bloco["ordem"].max()) if not df_bloco.empty else -1

        for _, row in df_bloco.iterrows():
            dia_semana = row["Dia"].weekday()
            col_base = colunas_dias[dia_semana]
            ordem = int(row["ordem"])

            linha_atual = linha_global_1 + ordem
            colar_range(nova_aba, linha_dados, linha_atual, 1, False) # Formatação/estilo

            # Preenche dados fixos (só se ainda não tiver nada na linha)
            if not nova_aba.cell(row=linha_atual, column=1).value:
                nova_aba.cell(row=linha_atual, column=1).value = row["Codigo"]
                nova_aba.cell(row=linha_atual, column=2).value = row["Linha"]
                nova_aba.cell(row=linha_atual, column=3).value = MAPA_MODAL.get(row["Modal"], row["Modal"])
                nova_aba.cell(row=linha_atual, column=4).value = row["Sent"]

            # Horários
            if row["Observacao"] == "OK":
                nova_aba.cell(row=linha_atual, column=col_base).value = row["THor"]
                nova_aba.cell(row=linha_atual, column=col_base + 1).value = row["Passag"]
                nova_aba.cell(row=linha_atual, column=col_base + 2).value = row["Oferta"]
            else:
                nova_aba.cell(row=linha_atual, column=col_base + 3).value = row["THor"]
                nova_aba.cell(row=linha_atual, column=col_base + 4).value = row["Passag"]
                nova_aba.cell(row=linha_atual, column=col_base + 5).value = row["Oferta"]

        # depois de preencher todas as viagens desse código, avança o bloco de linhas para o próximo código
        linha_global_1 = linha_global_1 + max_ordem + 1
        
        # Atualiza progresso
        contador += 1
        barra.progress(contador / total_blocos, f"Processando semana: {semana_num}  |  Sentido: 1  |  Linha: {codigo}")

    ultima_linha_sent1 = linha_global_1 - 1

    # === 5) Inserir totalizador de Sent 1 ===
    linha_totalizador_1 = ultima_linha_sent1 + 1
    colar_range(nova_aba, totalizador_tpl, linha_totalizador_1, 1)

    preencher_totalizador(
        nova_aba,
        linha_inicio=7,
        linha_fim=ultima_linha_sent1,
        linha_totalizador=linha_totalizador_1,
        colunas_dias=colunas_dias
    )

    linha_atual = linha_totalizador_1 + 1

    # === 6) BLOCO SENT 2 ===
    if not df_sent2.empty:
        # === 7) Inserir cabeçalho para Sent 2 ===
        linha_cabecalho_2 = linha_totalizador_1 + 1
        colar_range(nova_aba, cabecalho, linha_cabecalho_2, 1)

        linha_global_2 = linha_cabecalho_2 + 2   # primeira linha de dados de Sent 2
        inicio_sent2 = linha_global_2

        for (codigo, sent), df_bloco in df_sent2.groupby(["Codigo", "Sent"], sort=False):
            df_bloco = df_bloco.copy()
            df_bloco["ordem"] = df_bloco.groupby("Dia").cumcount()
            max_ordem = int(df_bloco["ordem"].max()) if not df_bloco.empty else -1

            for _, row in df_bloco.iterrows():
                dia_semana = row["Dia"].weekday()
                col_base = colunas_dias[dia_semana]
                ordem = int(row["ordem"])

                linha_atual = linha_global_2 + ordem
                colar_range(nova_aba, linha_dados, linha_atual, 1, False) # Formatação/estilo
                
                if not nova_aba.cell(row=linha_atual, column=1).value:
                    nova_aba.cell(row=linha_atual, column=1).value = row["Codigo"]
                    nova_aba.cell(row=linha_atual, column=2).value = row["Linha"]
                    nova_aba.cell(row=linha_atual, column=3).value = MAPA_MODAL.get(row["Modal"], row["Modal"])
                    nova_aba.cell(row=linha_atual, column=4).value = row["Sent"]

                if row["Observacao"] == "OK":
                    nova_aba.cell(row=linha_atual, column=col_base).value = row["THor"]
                    nova_aba.cell(row=linha_atual, column=col_base + 1).value = row["Passag"]
                    nova_aba.cell(row=linha_atual, column=col_base + 2).value = row["Oferta"]
                else:
                    nova_aba.cell(row=linha_atual, column=col_base + 3).value = row["THor"]
                    nova_aba.cell(row=linha_atual, column=col_base + 4).value = row["Passag"]
                    nova_aba.cell(row=linha_atual, column=col_base + 5).value = row["Oferta"]

            linha_global_2 = linha_global_2 + max_ordem + 1

            # Atualiza progresso
            contador += 1
            barra.progress(contador / total_blocos, f"Processando semana: {semana_num}  |  Sentido: 2  |  Linha: {codigo}")

        linha_fim_sent2 = linha_global_2 - 1
        linha_totalizador_2 = linha_fim_sent2 + 1

        colar_range(nova_aba, totalizador_tpl, linha_totalizador_2, 1)

        preencher_totalizador(
            nova_aba,
            linha_inicio=inicio_sent2,
            linha_fim=linha_fim_sent2,
            linha_totalizador=linha_totalizador_2,
            colunas_dias=colunas_dias
        )

        # === 8) Inserir legenda ===
        linha_legenda = linha_totalizador_2 + 4
        colar_range(nova_aba, legenda, linha_legenda, 1)
    else:
        # se não tiver Sent 2, legenda vem logo após totalizador 1
        linha_legenda = linha_totalizador_1 + 4
        colar_range(nova_aba, legenda, linha_legenda, 1)

# endregion

# Configuração da página
st.set_page_config(layout="wide")

# Layout
st.header("📄 [PDO] Dados Operacionais", anchor=False)
st.divider()

with st.container():
    col1, col2, col3, col4 = st.columns([2, 2, 2, 1], vertical_alignment='top')

    with col1:
        # Upload do arquivo de dados de passageiros
        st.subheader("Dados de passageiros", help="Transnet > Módulos > Tráfego/Arrecadação > Consultas/Relatórios > Controle Operacional/Tráfego > Desempenho Diário das Linhas", anchor=False)
        up_passageiros = st.file_uploader("Arquivo Relatório Desempenho Diário das Linhas.csv", type='csv', key=1)
        
    with col2:
        # Upload do arquivo dos dados das viagens
        st.subheader("Dados de viagens", help="Transnet > Módulos > Tráfego/Arrecadação > Consultas/Relatórios > Controle Operacional/Tráfego > Controle Operacional Detalhado Por Linha", anchor=False)
        up_viagens = st.file_uploader("Arquivo Relatório Controle Operacional Detalhado por Linha.csv", type='csv', key=2)

    with col3:
        # Upload da planilha para conferência das viagens
        st.subheader("Viagens Previstas", help="Planilha de viagens previstas", anchor=False)
        up_conferencia = st.file_uploader("Selecione um arquivo .XLSX", type='xlsx', key=3)

with st.container():       
    col1, col2, col3 = st.columns([3, 3, 3], vertical_alignment='top')
    with col1:
        # Feriados
        st.subheader("Feriados", anchor=False)

        df_feriado = pd.DataFrame([{"data": None, "escala": None}])

        # Editor de tabela
        df_feriado_editado = st.data_editor(
            df_feriado,
            num_rows="dynamic",
            column_config={
                "data": st.column_config.DateColumn("Data", format="DD/MM/YYYY"),
                "escala": st.column_config.SelectboxColumn("Escala", options=["Sábado", "Domingo"])
            }
        )

        # Converte depois do editor
        df_feriado_editado["data"] = pd.to_datetime(df_feriado_editado["data"], errors="coerce").dt.date

botao = st.sidebar.button("Iniciar", type="primary")

st.divider()
placeholder = st.empty() 

if botao:
    try:
        # Remove os botões
        st.session_state.pop("pdo", None)

        # Verificações de seleção dos arquivos
        if up_passageiros is None:
            placeholder.warning("Arquivo Relatório Desempenho Diário das Linhas não foi selecionado!", icon=":material/error_outline:")
            st.stop()

        if up_viagens is None:
            placeholder.warning("Arquivo Relatório Controle Operacional Detalhado por Linha!", icon=":material/error_outline:")        
            st.stop()

        if up_conferencia is None:
            placeholder.warning("Planilha para conferência não foi selecionada!", icon=":material/error_outline:")
            st.stop()

        with placeholder.status("Processando, aguarde...", expanded=True) as status:
            msg = st.empty()
            msg.write("📄 Lendo os arquivos...")
            config = json_utils.ler_json("config.json") # Arquivo de configuração
            df_linhas = pd.DataFrame(json_utils.ler_json(config["matrizes"]["linhas"])) # Matriz de linhas
            df_det1 = files_utils.ler_detalhado_linha(up_viagens) # Arquivo detalhado por linha
            df_prev_met = files_utils.ler_viagens_previstas(up_conferencia) # Planilha de viagens previstas Metroplan
            # Arquivo desempenho diário das linhas
            #‼️‼️‼️‼️‼️‼️‼️‼️

            
            msg.write("🧠 Processando - Controle Operacional Detalhado por Linha...")
            # Dropa colunas desnecessárias
            columns_to_drop = ['#', 'Orig', 'Dest', 'Dif', 'Parado', 'Prev', 'Real2', 'Dif2', 'CVg', 'Veiculo', 'Docmto', 'Motorista', 'Cobrador', 'EmPe', 'Oferta', 'Km_h', 'Meta', 'CVg2', 'TipoViagem']
            df_det1 = df_det1.drop(columns=columns_to_drop)
            # Renomeia capacidade sentada para Oferta
            df_det1.rename(columns={'Sent.1': 'Oferta'}, inplace=True)
            # Merge com arquivo de linhas para ter a modalidade/serviço
            df_det1 = df_det1.merge(df_linhas[["Cod_Met", "Modal"]], left_on="Codigo", right_on="Cod_Met", how="left")
            df_det1 = df_det1.drop(columns=["Cod_Met"])
            df_det1["Observacao"] = df_det1["Observacao"].astype(str).str.strip()
            # Exclui viagens que NÃO TEM passageiros (possíveis erros de digitação)
            df_det2 = df_det1[~(df_det1["Passag"].isna() & (df_det1["Observacao"] != "Furo de Viagem"))]
            # Preenche o horário das previstas com as realizadas nas viagens extras (vou usar esta coluna para os horários)
            df_det2.loc[:, "THor"] = df_det2["THor"].fillna(df_det2["Real"])
            # Converte e ordena
            df_det2["Dia"] = pd.to_datetime(df_det2["Dia"], dayfirst=True, errors="coerce").dt.date
            df_det2 = df_det2.sort_values(["Sent", "Codigo", "Dia", "THor"])

            msg.write("🧠 Processando - Desempenho Diário das Linhas...")
            #‼️‼️‼️‼️‼️‼️‼️‼️

            msg.write("🧠 Processando os dados para conferência...")
            df_conf1 = df_det2[["Codigo", "Sent", "Dia", "Observacao"]]
            df_conf1["Dia Semana"] = df_conf1["Dia"].map(lambda d: {0:"U",1:"U",2:"U",3:"U",4:"U",5:"S",6:"D"}[d.weekday()]) # Crio Dia Semana
            # Atualizo com o feriado, se houver
            df_feriado_editado = df_feriado_editado.dropna(how="all") # Limpa se não houver dados
            if not df_feriado_editado.empty:
                # 1. Faz o merge com base na data
                df_conf_merged = df_conf1.merge(df_feriado_editado, left_on='Dia', right_on='data', how='left')
                # 2. Atualiza 'Dia Semana' com base na escala
                df_conf_merged.loc[df_conf_merged['escala'] == 'Sábado', 'Dia Semana'] = 'S'
                df_conf_merged.loc[df_conf_merged['escala'] == 'Domingo', 'Dia Semana'] = 'D'
                # 3. Remove colunas extras
                df_conf1 = df_conf_merged

            df_conf1 = df_conf1.drop(columns=["Dia", "data", "escala"], errors="ignore") # Dropa colunas incluindo as usadas no feriado
            # Agrupa
            df_conf2 = (
                df_conf1
                    .assign(
                        Tipo=lambda x: x['Observacao'].map({
                            'OK': 'ERG',
                            'Viagem Extra': 'EXT',
                            'Furo de Viagem': 'FURO'
                        }),
                        Chave=lambda x: x['Tipo'] + '_' + x['Dia Semana'] + x['Sent'].astype(str)
                    )
                    .groupby(['Codigo', 'Chave'])
                    .size()
                    .unstack(fill_value=0)
                    .reindex(columns=['ERG_U1', 'EXT_U1', 'FURO_U1', 'ERG_U2', 'EXT_U2', 'FURO_U2',
                                    'ERG_S1', 'EXT_S1', 'FURO_S1', 'ERG_S2', 'EXT_S2', 'FURO_S2',
                                    'ERG_D1', 'EXT_D1', 'FURO_D1', 'ERG_D2', 'EXT_D2', 'FURO_D2'],
                                        fill_value=0)
                    .reset_index()
            )
            # Merge com previstas Metroplan
            df_conf2 = df_conf2.merge(
                df_prev_met,
                left_on="Codigo",
                right_on="Codigo",
                how="left"
            )

            # 🧩 Lendo Planilha modelo Modelo_PDO.xlsx
            msg.write("Lendo a planilha modelo...")
            wb_erg = load_workbook(config['pdo']['modelo_pdo'])
            wb_tm5 = load_workbook(config['pdo']['modelo_pdo'])

            # 1️⃣ CONFERÊNCIA
            msg.write("✏️ Editando a planilha ERG - Aba de conferência...")
            wb_erg = preencher_conferencia(wb_erg, df_conf2, False)
            msg.write("✏️ Editando a planilha TM5 - Aba de conferência...")
            wb_tm5 = preencher_conferencia(wb_tm5, df_conf2, True)
                        
            # 2️⃣ SEMANAS
            # Crio as abas das semanas e preencho os dias/feriados
            with msg.container():
                st.write("✏️ Editando a planilha ERG - Abas semanais...")
                barra = st.progress(0)
                wb_erg = criar_abas_com_dias(wb_erg, False)
            with msg.container():
                st.write("✏️ Editando a planilha TM5 - Abas semanais...")
                barra = st.progress(0)
                wb_tm5 = criar_abas_com_dias(wb_tm5, True)

            # 3️⃣ TOTAL GERAL
            msg.write("✏️ Editando a planilha ERG - Aba de totais...")
            wb_erg = preencher_totais(wb_erg)
            msg.write("✏️ Editando a planilha TM5 - Aba de totais...")
            wb_tm5 = preencher_totais(wb_tm5)

            # 🧩 Salvar em memória
            msg.write("💾 Salvando planilha ERG...")
            buffer_pdo_erg = BytesIO()
            wb_erg.save(buffer_pdo_erg)
            buffer_pdo_erg.seek(0)
            st.session_state["buffer_pdo_erg"] = buffer_pdo_erg # Arquivo ERG
            msg.write("💾 Salvando planilha TM5...")
            buffer_pdo_tm5 = BytesIO()
            wb_tm5.save(buffer_pdo_tm5)
            buffer_pdo_tm5.seek(0)
            st.session_state["buffer_pdo_tm5"] = buffer_pdo_tm5 # Arquivo TM5
            st.session_state["pdo"] = f"{df_det2.loc[0, "Dia"].strftime("%m.%Y")}" # Condição para os botões

            status.update(label="Concluído!", state="complete")

        placeholder.success("Arquivos gerados com sucesso!")

    except Exception as e:  
        placeholder.error(f"🐞 Erro: {traceback.format_exc()}")
    finally:
        df_linhas = None
        df_feriado_editado = None
        df_feriado = None
        df_det1 = None
        df_det2 = None
        df_conf1 = None
        df_conf2 = None
        df_conf_merged = None
        buffer_pdo = None
        wb_erg = None
        wb_tm5 = None

# ✳️ Downloads ✳️
if st.session_state.get("pdo", False):  
    col1, col2, col3 = st.columns([1,1,5], vertical_alignment='top')
    with col1:
        st.download_button(
            label="📥 Download PDO-ERG", 
            data=st.session_state["buffer_pdo_erg"], 
            file_name=f"GUAIBA [{st.session_state["pdo"]}].xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
        )
    with col2:     
        st.download_button(
            label="📥 Download PDO-TM5", 
            data=st.session_state["buffer_pdo_tm5"], 
            file_name=f"GUAIBA-TM5 [{st.session_state["pdo"]}].xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
        )